import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font

# 获取当前脚本所在的文件夹路径并提取文件夹名字
folder_name = os.path.basename(os.path.dirname(os.path.abspath(__file__)))

# 读取 SystemData.xlsx 文件
file_path = 'SystemData.xlsx'
df = pd.read_excel(file_path)

# 获取所有包含 'maxU'、'minU'、'maxT'、'totalChargeKwh'、'totalDischargeKwh' 的列
maxU_columns = [col for col in df.columns if 'maxU' in col]
minU_columns = [col for col in df.columns if 'minU' in col]
maxT_columns = [col for col in df.columns if 'maxT' in col]
totalChargeKwh_columns = [col for col in df.columns if 'totalChargeKwh' in col]
totalDischargeKwh_columns = [col for col in df.columns if 'totalDischargeKwh' in col]

# 创建新的列并计算
df['sysMaxU'] = df[maxU_columns].max(axis=1)
df['sysMinU'] = df[minU_columns].min(axis=1)
df['MaxDiff'] = (df['sysMaxU'] - df['sysMinU']) * 1000
df['sysMaxT'] = df[maxT_columns].max(axis=1)

# 计算 DayTotalChargeKwh 和 DayTotalDischargeKwh
df['DayTotalChargeKwh'] = sum(df[col].iloc[-1] - df[col].iloc[0] for col in totalChargeKwh_columns)
df['DayTotalDischargeKwh'] = sum(df[col].iloc[-1] - df[col].iloc[0] for col in totalDischargeKwh_columns)

# 按照需求将新增列插入到合适位置
# 插入 sysMaxU 到多个 maxU 列之后
for col in maxU_columns:
    col_index = df.columns.get_loc(col) + 1
    df.insert(col_index, 'sysMaxU', df.pop('sysMaxU'))

# 插入 sysMinU 到多个 minU 列之后
for col in minU_columns:
    col_index = df.columns.get_loc(col) + 1
    df.insert(col_index, 'sysMinU', df.pop('sysMinU'))

# 插入 MaxDiff 到 sysMinU 列之后
sysMinU_index = df.columns.get_loc('sysMinU') + 1
df.insert(sysMinU_index, 'MaxDiff', df.pop('MaxDiff'))

# 插入 sysMaxT 到多个 maxT 列之后
for col in maxT_columns:
    col_index = df.columns.get_loc(col) + 1
    df.insert(col_index, 'sysMaxT', df.pop('sysMaxT'))

# 插入 DayTotalChargeKwh 到多个 totalChargeKwh 列之后
for col in totalChargeKwh_columns:
    col_index = df.columns.get_loc(col) + 1
    df.insert(col_index, 'DayTotalChargeKwh', df.pop('DayTotalChargeKwh'))

# 插入 DayTotalDischargeKwh 到多个 totalDischargeKwh 列之后
for col in totalDischargeKwh_columns:
    col_index = df.columns.get_loc(col) + 1
    df.insert(col_index, 'DayTotalDischargeKwh', df.pop('DayTotalDischargeKwh'))

# 根据当前文件夹的名字来命名输出文件
output_file_name = f"{folder_name}.xlsx"
output_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), output_file_name)

# 保存数据到新的 Excel 文件
df.to_excel(output_file_path, index=False)

# 打开生成的 Excel 文件并进行格式调整
wb = load_workbook(output_file_path)
ws = wb.active

# 设置文字颜色为红色
red_font = Font(color="FF0000")

# 设置 sysMaxU, sysMinU, MaxDiff, sysMaxT, DayTotalChargeKwh, DayTotalDischargeKwh 列文字颜色
highlight_columns = ['sysMaxU', 'sysMinU', 'MaxDiff', 'sysMaxT', 'DayTotalChargeKwh', 'DayTotalDischargeKwh']
for col in highlight_columns:
    col_index = df.columns.get_loc(col) + 1  # openpyxl 是 1 基索引
    for row in range(2, len(df) + 2):  # 从第二行开始，因为第一行是标题
        cell = ws.cell(row=row, column=col_index)
        cell.font = red_font

# 保存文件
wb.save(output_file_path)

print(f"处理完成，结果已保存到 {output_file_path}")
